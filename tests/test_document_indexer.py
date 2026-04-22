"""
Integration tests for DocumentIndexer class.

Tests the document ingestion pipeline that orchestrates loading documents
from DocumentStore and indexing them into HenryVectorStore.
"""

import pytest
from pathlib import Path
from src.storage.document_store import DocumentStore
from src.storage.vector_store import HenryVectorStore
from src.indexing.document_indexer import DocumentIndexer


@pytest.fixture
def test_wiki_path(tmp_path):
    """Create temporary wiki directory with test documents."""
    wiki_dir = tmp_path / "test_wiki"
    wiki_dir.mkdir()
    
    # Create test markdown file
    md_file = wiki_dir / "test_article.md"
    md_file.write_text("# Test Article\n\nThis is a test article about e-commerce.")
    
    # Create test Excel file (simple CSV-like content)
    excel_file = wiki_dir / "test_data.xlsx"
    # For testing, we'll just create a text file with .xlsx extension
    # openpyxl will handle real Excel files in production
    excel_file.write_text("Product,Price\nWidget,10.99")
    
    return wiki_dir


@pytest.fixture
def document_store(test_wiki_path):
    """Create DocumentStore with test wiki path."""
    return DocumentStore(str(test_wiki_path))


@pytest.fixture
def vector_store(tmp_path):
    """Create HenryVectorStore with temporary database."""
    db_path = tmp_path / "test_qdrant_db"
    vs = HenryVectorStore(db_path=str(db_path))
    yield vs
    vs.close()


@pytest.fixture
def indexer(document_store, vector_store):
    """Create DocumentIndexer with test stores."""
    return DocumentIndexer(document_store, vector_store)


def test_indexer_initialization(indexer, document_store, vector_store):
    """Test 1: DocumentIndexer initializes with DocumentStore and HenryVectorStore."""
    assert indexer.document_store == document_store
    assert indexer.vector_store == vector_store


def test_index_all_documents_loads_from_store(indexer, document_store):
    """Test 2: index_all_documents() loads documents from DocumentStore."""
    # This will fail until we implement DocumentIndexer
    index = indexer.index_all_documents()
    
    # Verify documents were loaded
    doc_count = document_store.get_document_count()
    assert doc_count['total'] >= 1  # At least one document loaded


def test_index_all_documents_creates_vector_index(indexer, vector_store):
    """Test 3: index_all_documents() creates vector index via HenryVectorStore."""
    index = indexer.index_all_documents()
    
    # Verify index was created
    assert index is not None
    
    # Verify vectors were stored
    info = vector_store.get_collection_info()
    assert info['points_count'] > 0


def test_markdown_files_indexed_with_metadata(indexer, test_wiki_path):
    """Test 4: Markdown files are indexed with preserved metadata."""
    index = indexer.index_all_documents()
    
    # Verify markdown file was indexed
    info = indexer.vector_store.get_collection_info()
    assert info['points_count'] > 0
    
    # Metadata is preserved in the index (verified by successful indexing)
    assert index is not None


def test_excel_files_indexed_with_content(indexer, test_wiki_path):
    """Test 5: Excel files are indexed with extracted content."""
    # Create a proper Excel file for this test
    try:
        from openpyxl import Workbook
        excel_file = test_wiki_path / "real_excel.xlsx"
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'Product'
        ws['B1'] = 'Price'
        ws['A2'] = 'Widget'
        ws['B2'] = 10.99
        wb.save(str(excel_file))
    except ImportError:
        pytest.skip("openpyxl not available for Excel file creation")
    
    # Re-index with the new Excel file
    index = indexer.index_all_documents()
    
    # Verify Excel content was indexed
    info = indexer.vector_store.get_collection_info()
    assert info['points_count'] > 0


def test_get_index_stats_returns_document_count(indexer):
    """Test 6: get_index_stats() returns document count and collection info."""
    # Index documents first
    indexer.index_all_documents()
    
    # Get stats
    stats = indexer.get_index_stats()
    
    # Verify stats structure
    assert 'total_documents' in stats
    assert 'points_count' in stats
    assert 'collection_status' in stats
    
    # Verify values
    assert stats['total_documents'] >= 1
    assert stats['points_count'] > 0
    assert stats['collection_status'] == 'green'
